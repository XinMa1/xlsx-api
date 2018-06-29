#!/usr/bin/env python
# -*- coding: utf-8 -*-
import json
import itertools
from hashlib import md5
from pymongo import MongoClient
from pymongo.errors import DuplicateKeyError
from openpyxl import load_workbook

conn = MongoClient()
db = conn.xlsxs
collection_summary = db.xlsx_summary
collection_summary.create_index([("plat_name", 1)], unique=True)


class FileUpload(object):
    def __init__(self, xlsx_file, plat_name):
        self.xlsx_file = xlsx_file
        self.plat_name = plat_name
        if not db.get_collection(self.plat_name):
            collection = db.create_collection(self.plat_name)
            collection.create_index(
                [("hash", 1)],
                unique=True,
                background=True)
            self.collection = collection
        else:
            self.collection = db.get_collection(self.plat_name)
        summary = collection_summary.find_one({"plat_name": self.plat_name})
        self.sheets_name = summary["sheets_name"]
        self.col_dict = summary["col_dict"]

    # 上传总处理
    def upload(self):
        wb = load_workbook(self.xlsx_file)
        xlsx_count = 0
        for i in self.sheets_name:
            if i:
                ws = wb[i]
                self.ws = ws
                # 转换xlsx文件头部
                col_list = self.get_headers()
                # 验证header是否匹配
                code, msg = self.verify_headers(i)
                if code:
                    return code, msg
                nrows = ws.max_row
                if nrows - 1:
                    xlsx_count += (nrows - 1)
                    # 处理数据
                    xlsx_data = self.handle_data(col_list)
                    # 数据插入
                    self.insert_data(xlsx_data)

        if self.count_verify(xlsx_count):
            return 1, u'上传总条数和数据库总条数不匹配,\
                                       请重新核验数据(已清表)!'

        return 0, u'上传xlsx文件成功!'

    # 核验数据
    def verify_headers(self, i):
        ncols = self.ws.max_column
        summary = collection_summary.find_one({"plat_name": self.plat_name})
        if summary['cols'] != ncols:
            self.collection.remove()
            return 2, u"sheet: %s header 缺失! 数据表已清空,请核验数据后再做尝试" % i
        return 0, "success verify headers"

    # 处理数据
    def handle_data(self, col_list):
        nrows = self.ws.max_row
        ncols = self.ws.max_column
        xlsx_data = []
        for row in xrange(1, nrows + 1):
            if row != 1:
                json_value = {}
                for col in xrange(1, ncols + 1):
                    for index, value in enumerate(col_list, 1):
                        if col == index:
                            json_value[value] = self.ws.\
                                    cell(row=row, column=col).value
                xlsx_data.append(json_value)
        return xlsx_data

    def get_headers(self):
        ncols = self.ws.max_column
        col_list = []
        col_count = 0
        for col in xrange(1, ncols+1):
            cell_value = self.ws.cell(row=1, column=col).value
            for k, v in self.col_dict.items():
                if cell_value == k:
                    col_count += 1
                    col_list.append(v)
        return col_list

    # 条数校验
    def count_verify(self, max_rows):
        total_count = self.collection.find().count()

        if max_rows != total_count:
            self.collection.remove()
            return 1
        return 0

    # 插入数据
    def insert_data(self, items):
        '''
            xlsx data:
                type: list
        '''
        items = sorted(items, key=self.generate_hash)

        for k, g in itertools.groupby(items, key=self.generate_hash):
            g = list(g)
            for i, item in enumerate(g):
                item["extraid"] = str(i)

        for item in items:
            item["hash"] = self.generate_hash(item)

        n_inserted = 0
        for item in items:

            try:
                self.collection.insert_one(item)
                n_inserted += 1
            except DuplicateKeyError:
                pass
        print("n_inserted: %s" % n_inserted)

    # 生成hash值
    def generate_hash(self, item):
        # s = ','.join(itertools.chain(*sorted(item.items())))
        s = ','.join([str(v) for v in itertools.chain(*sorted(item.items()))
                      ])
        return md5(s).hexdigest()


class SummaryHandle(object):
    def __init__(self, plat_name, ncols, sheets_name, col_dict):
        self.plat_name = plat_name
        self.ncols = ncols
        self.sheets_name = sheets_name
        self.col_dict = col_dict

    def summary_list(self):
        summary_list = collection_summary.find()
        return json.dumps(summary_list)

    def create_or_modify(self):
        data = {"plat_name": self.plat_name, "cols": self.ncols, "col_dict":
                self.col_dict, "sheets_name": self.sheets_name}
        try:
            collection_summary.save(data)
        except DuplicateKeyError:
            pass

        return u"创建或修改成功!"


